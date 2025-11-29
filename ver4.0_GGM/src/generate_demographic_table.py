import pandas as pd
import numpy as np
import re
from scipy import stats
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from pathlib import Path
import os

SHORT_TERM = {
    'Hypertension': 'HTN',
    'Diabetes Mellitus': 'DM',
    'Hyperlipidemia': 'HL',
    'Angina/Myocardial Infarction': 'AMI',
    'Stroke': 'Stroke',
    'Chronic kidney disease (eGFR<60)': 'CKD',
    'Obese': 'Obese',
    'Increased waist circumference': '△ WC',
    'Elevated blood pressure': '△ BP',
    'Impaired fasting glucose': 'IFG',
    'Elevated triglycerides': '△ TG',
    'Decreased HDL-C': '▽ HDL-C',
    'Systolic blood pressure (mmHg)': 'SBP',
    'Diastolic Blood Pressure (mmHg)': 'DBP',
    'Fasting glucose (mg/dL)': 'FBS',
    'Triglycerides (mg/dL)': 'TG',
    'HbA1c (%)': 'HBA1C',
    'HDL-C (mg/dL)': 'HDL-C.',
    'LDL-C (mg/dL)': 'LDL-C.',
    'eGFR (mL/min/1.73m2)': 'eGFR',
    'Weight (kg)': 'Weight',
    'BMI(kg/m2)': 'BMI',
    'Waist circumference (cm)': 'WC',
    'Physical activity': 'PA',
    'Alcohol Consumption': 'Alcohol',
    'Current smoking': 'Smoking',
    'Grain Products': 'Grain',
    'Protein Foods': 'Protein',
    'Vegetables': 'Vegetables',
    'Dairy Products': 'Dairy',
    'Fruits': 'Fruits',
    'Fried Foods': 'Fried',
    'High Fat Meat': 'High Fat\nMeat',
    'Processed Foods': 'Processed\nFoods',
    'Sugar-Sweetened Beverages': 'SSB',
    'Additional Salt Use': 'Add-Salt',
    'Salty Food Consumption': 'Salty\nFood',
    'Sweet Food Consumption': 'Sweet\nFood',
}

REVERSE_MAPPING = {
    '△ WC': 'Increased waist circumference',
    '△ BP': 'Elevated blood pressure',
    'IFG': 'Impaired fasting glucose',
    '△ TG': 'Elevated triglycerides',
    '▽ HDL-C': 'Decreased HDL-C',
    'Weight': 'Weight (kg)',
    'BMI': 'BMI(kg/m2)',
    'WC': 'Waist circumference (cm)',
    'HTN': 'Hypertension',
    'DM': 'Diabetes Mellitus',
    'HL': 'Hyperlipidemia',
    'AMI': 'Angina/Myocardial Infarction',
    'CKD': 'Chronic kidney disease (eGFR<60)',
    'SBP': 'Systolic blood pressure (mmHg)',
    'DBP': 'Diastolic Blood Pressure (mmHg)',
    'TG': 'Triglycerides (mg/dL)',
    'LDL-C.': 'LDL-C (mg/dL)',
    'HDL-C.': 'HDL-C (mg/dL)',
    'FBS': 'Fasting glucose (mg/dL)',
    'HBA1C': 'HbA1c (%)',
    'eGFR': 'eGFR (mL/min/1.73m2)',
    'PA': 'Physical activity',
    'Alcohol': 'Alcohol Consumption',
    'Smoking': 'Current smoking',
    'Grain': 'Grain Products',
    'Protein': 'Protein Foods',
    'Dairy': 'Dairy Products',
    'Fried': 'Fried Foods',
    'High Fat\nMeat': 'High Fat Meat',
    'Processed\nFoods': 'Processed Foods',
    'SSB': 'Sugar-Sweetened Beverages',
    'Add-Salt': 'Additional Salt Use',
    'Salty\nFood': 'Salty Food Consumption',
    'Sweet\nFood': 'Sweet Food Consumption'
}

AGE_COL = ['Age']
AM_COL = ['Height (cm)', 'Weight (kg)', 'BMI(kg/m2)', 'Waist circumference (cm)']
CM_COL = ['Systolic blood pressure (mmHg)', 'Diastolic Blood Pressure (mmHg)', 'Triglycerides (mg/dL)',
          'HDL-C (mg/dL)', 'LDL-C (mg/dL)', 'Non-HDL-C (mg/dL)', 'Total Cholesterol (mg/dL)',
          'Fasting glucose (mg/dL)', 'HbA1c (%)', 'eGFR (mL/min/1.73m2)']
METS_COLS = ['Increased waist circumference', 'Elevated blood pressure', 'Impaired fasting glucose',
             'Elevated triglycerides', 'Decreased HDL-C']
DISEASE_COLS = ['Hypertension', 'Antihypertensive Medication', 'Hyperlipidemia', 'Statin Medication']

CONTINUOUS_COLS = AGE_COL + AM_COL + CM_COL

def clean_and_convert(x):
    if isinstance(x, str):
        match = re.search(r'[-+]?\d*\.?\d+', x)
        if match:
            return float(match.group())
    return x

def format_numeric(values):
    if len(values) > 0:
        return f"{values.mean():.2f}±{values.std():.2f}"
    return ""

def format_numeric2(values):
    if len(values) > 0:
        median = np.median(values)
        q1 = np.percentile(values, 25)
        q3 = np.percentile(values, 75)
        return f"{median:.2f} ({q1:.2f}-{q3:.2f})"
    return ""

def format_categorical(count, total):
    if total > 0:
        return f"{count} ({count/total*100:.1f}%)"
    return "0 (0.0%)"

def age_filter(df, age_range):
    if age_range == "under40":
        return df[df['Age'] < 40]
    elif age_range == "40-59":
        return df[(df['Age'] >= 40) & (df['Age'] < 60)]
    elif age_range == "60up":
        return df[df['Age'] >= 60]
    else:
        return df

def get_category(col):
    """Map variables to their display categories in the specified order"""
    category_mapping = {
        'Age': 'Age',
        'Sex': 'Male',
        'Education Level': 'Education Level',
        'Marital Status': 'Marital Status',
        'Household Income': 'Household Income',
        'BMI category': 'BMI category',
        'Increased waist circumference': 'WC (M>=90, F>=85)',
        'Height (cm)': 'Anthropometric Measures',
        'Weight (kg)': 'Anthropometric Measures',
        'BMI(kg/m2)': 'Anthropometric Measures',
        'Waist circumference (cm)': 'Anthropometric Measures',
        'Hypertension': 'Disease status & Medication',
        'Antihypertensive Medication': 'Disease status & Medication',
        'Hyperlipidemia': 'Disease status & Medication',
        'Statin Medication': 'Disease status & Medication',
        'Systolic blood pressure (mmHg)': 'Clinical Measures',
        'Diastolic Blood Pressure (mmHg)': 'Clinical Measures',
        'Total Cholesterol (mg/dL)': 'Clinical Measures',
        'Triglycerides (mg/dL)': 'Clinical Measures',
        'LDL-C (mg/dL)': 'Clinical Measures',
        'HDL-C (mg/dL)': 'Clinical Measures',
        'Non-HDL-C (mg/dL)': 'Clinical Measures',
        'Fasting glucose (mg/dL)': 'Clinical Measures',
        'HbA1c (%)': 'Clinical Measures',
        'eGFR (mL/min/1.73m2)': 'Clinical Measures',
        'Obese': 'Obese',
        'Physical activity': 'Physical activity',
        'Alcohol Consumption': 'Alcohol Consumption',
        'Current smoking': 'Current smoking',
        'Grain Products': 'Grain Products',
        'Protein Foods': 'Protein Foods',
        'Vegetables': 'Vegetables',
        'Dairy Products': 'Dairy Products',
        'Fruits': 'Fruits',
        'Fried Foods': 'Fried Foods',
        'High Fat Meat': 'High Fat Meat',
        'Processed Foods': 'Processed Foods',
        'Sugar-Sweetened Beverages': 'Sugar-Sweetened Beverages',
        'Additional Salt Use': 'Additional Salt Use',
        'Salty Food Consumption': 'Salty Food Consumption',
        'Sweet Food Consumption': 'Sweet Food Consumption',
        'Elevated blood pressure': 'Metabolic syndrome',
        'Impaired fasting glucose': 'Metabolic syndrome',
        'Elevated triglycerides': 'Metabolic syndrome',
        'Decreased HDL-C': 'Metabolic syndrome',
        'MetS': 'Metabolic syndrome'
    }

    return category_mapping.get(col, col)

def sep_group(base_df):
    return {
        "Total": base_df,
        "MetS": base_df[base_df['MetS'] == 1],
        "Non-MetS": base_df[base_df['MetS'] == 0]
    }

def read_p(p_value):
    if pd.isna(p_value):
        return ""
    elif p_value == "-":
        return "-"
    else:
        return f"{p_value:.3f}"


def move_column(ws, old_index, new_index):
    max_row = ws.max_row

    column_data = []
    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=old_index)
        column_data.append(cell.value)

    if old_index < new_index:
        for row in range(1, max_row + 1):
            for col in range(old_index, new_index):
                ws.cell(row=row, column=col).value = ws.cell(row=row, column=col+1).value
    else:
        for row in range(1, max_row + 1):
            for col in range(old_index, new_index, -1):
                ws.cell(row=row, column=col).value = ws.cell(row=row, column=col-1).value

    for row in range(1, max_row + 1):
        ws.cell(row=row, column=new_index).value = column_data[row-1]

def load_and_prepare_data(file_path):
    df = pd.read_csv(file_path)
    df.rename(columns=SHORT_TERM, inplace=True)
    df.rename(columns=REVERSE_MAPPING, inplace=True)
    return df

def preprocess_data(df):
    for col in METS_COLS:
        if col in df.columns:
            df[col] = df[col].astype(int)

    # Reorder columns according to the specified order
    df = df[[
        'R-ID', 'Age', 'Sex', 'Education Level', 'Marital Status', 'Household Income',
        'BMI category', 'Increased waist circumference',
        'Height (cm)', 'Weight (kg)', 'BMI(kg/m2)', 'Waist circumference (cm)',
        'Hypertension', 'Antihypertensive Medication', 'Hyperlipidemia', 'Statin Medication',
        'Systolic blood pressure (mmHg)', 'Diastolic Blood Pressure (mmHg)',
        'Total Cholesterol (mg/dL)', 'Triglycerides (mg/dL)', 'LDL-C (mg/dL)', 'HDL-C (mg/dL)',
        'Non-HDL-C (mg/dL)', 'Fasting glucose (mg/dL)', 'HbA1c (%)', 'eGFR (mL/min/1.73m2)',
        'Obese', 'Physical activity', 'Alcohol Consumption', 'Current smoking',
        'Grain Products', 'Protein Foods', 'Vegetables', 'Dairy Products', 'Fruits',
        'Fried Foods', 'High Fat Meat', 'Processed Foods', 'Sugar-Sweetened Beverages',
        'Additional Salt Use', 'Salty Food Consumption', 'Sweet Food Consumption',
        'Elevated blood pressure', 'Impaired fasting glucose',
        'Elevated triglycerides', 'Decreased HDL-C', 'MetS'
    ]].copy()

    df[CONTINUOUS_COLS] = df[CONTINUOUS_COLS].replace('Missing value', np.nan).fillna(0).astype(float)

    all_df = df.copy()
    men_df = df[df['Sex'] == 'M'].copy()
    women_df = df[df['Sex'] == 'F'].copy()

    return all_df, men_df, women_df

def calculate_p_value_two_groups(group1, group2, col, categorical_cols):
    values1 = group1[col].apply(clean_and_convert).dropna()
    values2 = group2[col].apply(clean_and_convert).dropna()

    if col in CONTINUOUS_COLS:
        _, p_value = stats.ttest_ind(values1, values2, equal_var=False)
        return p_value

    elif col in categorical_cols:
        values1_str = values1.astype(str)
        values2_str = values2.astype(str)

        all_unique = np.unique(np.concatenate([values1_str, values2_str]))

        if set(all_unique).issubset({'0', '1'}):
            cont_table = pd.crosstab(
                pd.Series(np.concatenate([values1_str, values2_str])),
                pd.Series(np.concatenate([np.ones(len(values1_str)), np.zeros(len(values2_str))]))
            )
            for category in ['0', '1']:
                if category not in cont_table.index:
                    cont_table.loc[category] = [0, 0]
            cont_table = cont_table.fillna(0)
            if cont_table.shape[0] > 1 and cont_table.shape[1] > 1:
                _, p_value, _, _ = stats.chi2_contingency(cont_table)
                return p_value

        else:
            contingency = {}
            for category in all_unique:
                contingency[category] = [sum(values1_str == category), sum(values2_str == category)]
            cont_table = pd.DataFrame.from_dict(contingency, orient='index')
            if cont_table.shape[0] > 1 and not (cont_table == 0).any(axis=None):
                _, p_value, _, _ = stats.chi2_contingency(cont_table)
                return p_value

    return np.nan

def calculate_12groups_p_value(df, col, categorical_cols):
    groups_data = []
    groups_labels = []

    for gender in ['M', 'F']:
        for age_range in ["under40", "40-59", "60up"]:
            for mets_status in [0, 1]:
                gender_df = df[df['Sex'] == gender]
                age_df = age_filter(gender_df, age_range)
                group_df = age_df[age_df['MetS'] == mets_status]

                values = group_df[col].apply(clean_and_convert).dropna()

                if col in CONTINUOUS_COLS:
                    groups_data.append(list(values))
                else:
                    groups_data.extend(values)
                    group_label = f"{gender}_{age_range}_MetS{mets_status}"
                    groups_labels.extend([group_label] * len(values))

    if col in CONTINUOUS_COLS:
        groups_data = [g for g in groups_data if len(g) > 0]
        if len(groups_data) >= 2:
            try:
                _, p_value = stats.f_oneway(*groups_data)
                return p_value
            except:
                return np.nan
        return np.nan

    elif col in categorical_cols:
        if len(groups_data) == 0:
            return np.nan

        groups_labels_array = np.array(groups_labels)
        groups_data_array = np.array(groups_data).astype(str)

        unique_groups = np.unique(groups_labels_array)
        unique_values = np.unique(groups_data_array)

        contingency_table = np.zeros((len(unique_values), len(unique_groups)))
        for i, val in enumerate(unique_values):
            for j, group in enumerate(unique_groups):
                contingency_table[i, j] = np.sum((groups_data_array == val) & (groups_labels_array == group))

        if contingency_table.shape[0] > 1 and contingency_table.shape[1] > 1:
            try:
                _, p_value, _, _ = stats.chi2_contingency(contingency_table)
                return p_value
            except:
                return np.nan
        return np.nan

    return np.nan

def calculate_age_groups_p_value(df, col, categorical_cols):
    age_groups = ["under40", "40-59", "60up"]
    age_data = []
    age_labels = []

    for age_range in age_groups:
        age_df = age_filter(df, age_range)
        values = age_df[col].apply(clean_and_convert).dropna()
        age_data.extend(values)
        age_labels.extend([age_range] * len(values))

    if col in CONTINUOUS_COLS:
        groups = []
        for age_range in age_groups:
            group_data = [age_data[i] for i in range(len(age_data)) if age_labels[i] == age_range]
            groups.append(group_data)
        try:
            _, p_value = stats.f_oneway(*groups)
            return p_value
        except:
            return np.nan

    elif col in categorical_cols:
        age_labels_array = np.array(age_labels)
        age_data_array = np.array(age_data)
        unique_ages = np.unique(age_labels_array)
        unique_values = np.unique(age_data_array)

        contingency_table = np.zeros((len(unique_values), len(unique_ages)))
        for i, val in enumerate(unique_values):
            for j, age in enumerate(unique_ages):
                contingency_table[i, j] = np.sum((age_data_array == val) & (age_labels_array == age))

        if contingency_table.shape[0] > 1 and contingency_table.shape[1] > 1:
            try:
                _, p_value, _, _ = stats.chi2_contingency(contingency_table)
                return p_value
            except:
                return '-'
        else:
            return '-'

    return '-'


def generate_summary_all_subjects(df, categorical_cols):
    mets_groups = sep_group(df)
    categorized_data = {}

    for col in list(df.columns[1:]):
        category = get_category(col)

        if category not in categorized_data:
            categorized_data[category] = []
            category_row = {'Variable': category}
            for mets_status in mets_groups.keys():
                category_row[mets_status] = category
            categorized_data[category].append(category_row)

        if col in CONTINUOUS_COLS:
            row = {'Variable': col}
            for mets_status, mets_df in mets_groups.items():
                values = mets_df[col].apply(clean_and_convert).dropna()
                if col == 'Triglycerides (mg/dL)':
                    row[mets_status] = format_numeric2(values)
                else:
                    row[mets_status] = format_numeric(values)
            categorized_data[category].append(row)

        else:
            temp_data = {mets_status: mets_df[col].astype(str).replace('nan', 'Missing value').value_counts()
                         for mets_status, mets_df in mets_groups.items()}
            base_counts = df[col].astype(str).replace('nan', 'Missing value').value_counts()
            categories = list(base_counts.keys()) if len(base_counts) > 0 else []

            binary_numeric = set(categories).issubset({'0', '1', 'Missing value'})
            binary_boolean = set(categories).issubset({'True', 'False', 'Missing value'})

            group_counts = {mets_status: len(mets_df) for mets_status, mets_df in mets_groups.items()}

            if col == 'Sex':
                row = {'Variable': 'Male'}
                for mets_status, counts in temp_data.items():
                    male_count = counts.get('M', 0)
                    row[mets_status] = format_categorical(male_count, group_counts[mets_status])
                categorized_data[category].append(row)

            elif binary_numeric or binary_boolean:
                target_value = '1' if binary_numeric else 'True'
                row = {'Variable': col}

                for mets_status, counts in temp_data.items():
                    target_count = counts.get(target_value, 0)
                    row[mets_status] = format_categorical(target_count, group_counts[mets_status])
                categorized_data[category].append(row)

            else:
                cat_list = []
                missing_count = 0
                for val, count in base_counts.items():
                    if val == 'Missing value':
                        missing_count = count
                    else:
                        try:
                            cat_list.append((float(val), val, count))
                        except:
                            cat_list.append((-float('inf'), val, count))

                cat_list.sort(reverse=True)
                sorted_items = [(cat[1], cat[2]) for cat in cat_list]
                if missing_count > 0:
                    sorted_items.append(('Missing value', missing_count))

                for val, _ in sorted_items:
                    row = {'Variable': val}
                    for mets_status, counts in temp_data.items():
                        val_count = counts.get(val, 0)
                        row[mets_status] = format_categorical(val_count, group_counts[mets_status])
                    categorized_data[category].append(row)

    return categorized_data

def generate_summary_by_gender_age(df, categorical_cols):
    mets_groups = sep_group(df)
    categorized_data = {}

    for col in list(df.columns[1:]):
        category = get_category(col)

        if category not in categorized_data:
            categorized_data[category] = []
            category_row = {'Variable': category}
            for mets_status in mets_groups.keys():
                for age_range in ["under40", "40-59", "60up"]:
                    col_key = f"{mets_status}_{age_range}"
                    category_row[col_key] = category
            categorized_data[category].append(category_row)

        if col in CONTINUOUS_COLS:
            row = {'Variable': col}
            for mets_status, mets_df in mets_groups.items():
                for age_range in ["under40", "40-59", "60up"]:
                    age_df = age_filter(mets_df, age_range)
                    values = age_df[col].apply(clean_and_convert).dropna()
                    col_key = f"{mets_status}_{age_range}"
                    if col == 'Triglycerides (mg/dL)':
                        row[col_key] = format_numeric2(values)
                    else:
                        row[col_key] = format_numeric(values)
            categorized_data[category].append(row)

        else:
            temp_data = {}
            for mets_status, mets_df in mets_groups.items():
                for age_range in ["under40", "40-59", "60up"]:
                    age_df = age_filter(mets_df, age_range)
                    col_key = f"{mets_status}_{age_range}"
                    temp_data[col_key] = age_df[col].astype(str).replace('nan', 'Missing value').value_counts()

            base_counts = df[col].astype(str).replace('nan', 'Missing value').value_counts()
            categories = list(base_counts.keys()) if len(base_counts) > 0 else []
            binary_numeric = set(categories).issubset({'0', '1', 'Missing value'})
            binary_boolean = set(categories).issubset({'True', 'False', 'Missing value'})

            group_counts = {}
            for mets_status, mets_df in mets_groups.items():
                for age_range in ["under40", "40-59", "60up"]:
                    age_df = age_filter(mets_df, age_range)
                    col_key = f"{mets_status}_{age_range}"
                    group_counts[col_key] = len(age_df)

            if col == 'Sex':
                row = {'Variable': 'Male'}
                for col_key, counts in temp_data.items():
                    male_count = counts.get('M', 0)
                    row[col_key] = format_categorical(male_count, group_counts[col_key])
                categorized_data[category].append(row)

            elif binary_numeric or binary_boolean:
                target_value = '1' if binary_numeric else 'True'
                row = {'Variable': col}

                for col_key, counts in temp_data.items():
                    target_count = counts.get(target_value, 0)
                    row[col_key] = format_categorical(target_count, group_counts[col_key])
                categorized_data[category].append(row)

            else:
                cat_list = []
                missing_count = 0

                for val, count in base_counts.items():
                    if val == 'Missing value':
                        missing_count = count
                    else:
                        try:
                            cat_list.append((float(val), val, count))
                        except:
                            cat_list.append((-float('inf'), val, count))

                cat_list.sort(reverse=True)
                sorted_items = [(cat[1], cat[2]) for cat in cat_list]
                if missing_count > 0:
                    sorted_items.append(('Missing value', missing_count))

                for val, _ in sorted_items:
                    row = {'Variable': val}
                    for col_key, counts in temp_data.items():
                        val_count = counts.get(val, 0)
                        row[col_key] = format_categorical(val_count, group_counts[col_key])
                    categorized_data[category].append(row)

    return categorized_data

def generate_summary_12_groups(df, categorical_cols):
    """Generate summary statistics for 12 groups (Gender × Age × MetS)"""
    categorized_data = {}

    for col in list(df.columns[1:]):
        category = get_category(col)

        # Initialize category if not exists
        if category not in categorized_data:
            categorized_data[category] = []
            category_row = {'Variable': category}

            for gender in ['M', 'F']:
                for age_range in ["under40", "40-59", "60up"]:
                    for mets_status in [0, 1]:
                        col_key = f"{gender}_{age_range}_MetS{mets_status}"
                        category_row[col_key] = category
            categorized_data[category].append(category_row)

        # Handle continuous variables
        if col in CONTINUOUS_COLS:
            row = {'Variable': col}
            for gender in ['M', 'F']:
                gender_df = df[df['Sex'] == gender]
                for age_range in ["under40", "40-59", "60up"]:
                    age_df = age_filter(gender_df, age_range)
                    for mets_status in [0, 1]:
                        group_df = age_df[age_df['MetS'] == mets_status]
                        values = group_df[col].apply(clean_and_convert).dropna()
                        col_key = f"{gender}_{age_range}_MetS{mets_status}"
                        if col == 'Triglycerides (mg/dL)':
                            row[col_key] = format_numeric2(values)
                        else:
                            row[col_key] = format_numeric(values)
            categorized_data[category].append(row)

        # Handle categorical variables
        else:
            temp_data = {}
            group_counts = {}

            for gender in ['M', 'F']:
                gender_df = df[df['Sex'] == gender]
                for age_range in ["under40", "40-59", "60up"]:
                    age_df = age_filter(gender_df, age_range)
                    for mets_status in [0, 1]:
                        group_df = age_df[age_df['MetS'] == mets_status]
                        col_key = f"{gender}_{age_range}_MetS{mets_status}"
                        temp_data[col_key] = group_df[col].astype(str).replace('nan', 'Missing value').value_counts()
                        group_counts[col_key] = len(group_df)

            base_counts = df[col].astype(str).replace('nan', 'Missing value').value_counts()
            categories = list(base_counts.keys()) if len(base_counts) > 0 else []
            binary_numeric = set(categories).issubset({'0', '1', 'Missing value'})
            binary_boolean = set(categories).issubset({'True', 'False', 'Missing value'})

            # Special handling for Sex variable
            if col == 'Sex':
                row = {'Variable': 'Male'}
                for col_key, counts in temp_data.items():
                    male_count = counts.get('M', 0)
                    row[col_key] = format_categorical(male_count, group_counts[col_key])
                categorized_data[category].append(row)

            # Binary variables
            elif binary_numeric or binary_boolean:
                target_value = '1' if binary_numeric else 'True'
                row = {'Variable': col}

                for col_key, counts in temp_data.items():
                    target_count = counts.get(target_value, 0)
                    row[col_key] = format_categorical(target_count, group_counts[col_key])
                categorized_data[category].append(row)

            # Multi-category variables
            else:
                cat_list = []
                missing_count = 0

                for val, count in base_counts.items():
                    if val == 'Missing value':
                        missing_count = count
                    else:
                        try:
                            cat_list.append((float(val), val, count))
                        except:
                            cat_list.append((-float('inf'), val, count))

                cat_list.sort(reverse=True)
                sorted_items = [(cat[1], cat[2]) for cat in cat_list]
                if missing_count > 0:
                    sorted_items.append(('Missing value', missing_count))

                for val, _ in sorted_items:
                    row = {'Variable': val}
                    for col_key, counts in temp_data.items():
                        val_count = counts.get(val, 0)
                        row[col_key] = format_categorical(val_count, group_counts[col_key])
                    categorized_data[category].append(row)

    return categorized_data


def create_12groups_sheet(wb, results_by_category, twelve_groups_p_values, df, sheet_name='Table1'):
    """Create Excel sheet for 12 groups analysis"""
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Create headers - Row 1
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    cell = ws.cell(row=1, column=1, value="Variable")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    cell.font = bold_font

    # Gender headers (row 1)
    col_idx = 2
    for gender in ['Men', 'Women']:
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+5)
        cell = ws.cell(row=1, column=col_idx, value=gender)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.font = bold_font
        col_idx += 6

    # P-value header
    ws.merge_cells(start_row=1, start_column=14, end_row=3, end_column=14)
    cell = ws.cell(row=1, column=14, value="p-value\n(12 groups)")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    cell.font = bold_font

    # Age group headers (row 2)
    col_idx = 2
    for gender in ['M', 'F']:
        for age_label in ["age < 40", "40 ≤ age < 60", "60 ≤"]:
            ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx+1)
            cell = ws.cell(row=2, column=col_idx, value=age_label)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            cell.font = bold_font
            col_idx += 2

    # MetS status headers (row 3)
    col_idx = 2
    for gender in ['M', 'F']:
        for age_range in ["under40", "40-59", "60up"]:
            for mets_label in ["Non-MetS", "MetS"]:
                cell = ws.cell(row=3, column=col_idx, value=mets_label)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                cell.font = Font(bold=True, size=9)
                col_idx += 1

    # Add Total row (row 4) with group sizes
    row_idx = 4
    ws.cell(row=row_idx, column=1, value="Total").border = thin_border
    ws.cell(row=row_idx, column=1).font = bold_font

    col_idx = 2
    for gender in ['M', 'F']:
        gender_df = df[df['Sex'] == gender]
        for age_range in ["under40", "40-59", "60up"]:
            age_df = age_filter(gender_df, age_range)
            for mets_status in [0, 1]:
                group_df = age_df[age_df['MetS'] == mets_status]
                group_size = len(group_df)
                cell = ws.cell(row=row_idx, column=col_idx, value=group_size)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                col_idx += 1

    ws.cell(row=row_idx, column=14, value="").border = thin_border
    row_idx += 1

    p12_value_dict = dict(zip(twelve_groups_p_values['Variable'], twelve_groups_p_values['p_value']))

    for category, rows in results_by_category["12_Groups"].items():
        # Category header
        cell = ws.cell(row=row_idx, column=1, value=category)
        cell.font = bold_font
        cell.border = thin_border

        # P-value for category
        p12_value = p12_value_dict.get(category, np.nan)
        p12_value_str = read_p(p12_value)
        cell = ws.cell(row=row_idx, column=14, value=p12_value_str)
        cell.border = thin_border

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=13)
        row_idx += 1

        # Variable rows
        for i, row_data in enumerate(rows):
            # Skip redundant category row
            if i == 0:
                is_category_row = True
                for gender in ['M', 'F']:
                    for age_range in ["under40", "40-59", "60up"]:
                        for mets_status in [0, 1]:
                            col_key = f"{gender}_{age_range}_MetS{mets_status}"
                            if row_data.get(col_key, "") != category:
                                is_category_row = False
                                break
                        if not is_category_row:
                            break
                    if not is_category_row:
                        break

                if is_category_row:
                    continue

            variable_name = row_data['Variable']
            ws.cell(row=row_idx, column=1, value=variable_name).border = thin_border

            # Add data for 12 groups
            col_idx = 2
            for gender in ['M', 'F']:
                for age_range in ["under40", "40-59", "60up"]:
                    for mets_status in [0, 1]:
                        col_key = f"{gender}_{age_range}_MetS{mets_status}"
                        cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_key, ""))
                        cell.border = thin_border
                        col_idx += 1

            # P-value
            p12_value = p12_value_dict.get(variable_name, np.nan)
            p12_value_str = read_p(p12_value)
            cell = ws.cell(row=row_idx, column=14, value=p12_value_str)
            cell.border = thin_border

            row_idx += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 38
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col_letter].width = 15


def create_gender_age_sheet(wb, category_name, results_by_category, p_values_df):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)

    if category_name in wb.sheetnames:
        del wb[category_name]
    ws = wb.create_sheet(category_name)

    # Create headers
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    cell = ws.cell(row=1, column=1, value="Variable")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    cell.font = bold_font

    # MetS status headers
    for col, status in [(2, "Total"), (5, "MetS"), (8, "Non-MetS")]:
        cell = ws.cell(row=1, column=col, value=status)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.font = bold_font

    # P-value header
    cell = ws.cell(row=1, column=11, value="")
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    cell.font = bold_font

    # P-value sub-headers
    for i in range(3):
        col = 11 + i
        cell = ws.cell(row=2, column=col, value="P-value")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.font = bold_font

    # Age group headers
    age_labels = ["age < 40", "40 ≤ age < 60", "60 ≤"]
    for base_col in [2, 5, 8]:
        for i, age_label in enumerate(age_labels):
            cell = ws.cell(row=2, column=base_col + i, value=age_label)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            cell.font = bold_font

    # Add data rows
    row_idx = 3

    for category, rows in results_by_category[category_name].items():
        # Category header
        cell = ws.cell(row=row_idx, column=1, value=category)
        cell.font = bold_font
        cell.border = thin_border

        # P-values for category
        if category in p_values_df.index:
            for i, status in enumerate(['Total', 'MetS', 'NonMetS']):
                p_value = p_values_df.loc[category].get(status, np.nan)
                p_value_str = read_p(p_value)
                cell = ws.cell(row=row_idx, column=11 + i, value=p_value_str)
                cell.border = thin_border

        row_idx += 1

        # Variable rows
        for i, row_data in enumerate(rows):
            # Skip redundant category row
            if i == 0:
                is_category_row = True
                for mets_status in ["Total", "MetS", "Non-MetS"]:
                    for age_range in ["under40", "40-59", "60up"]:
                        col_key = f"{mets_status}_{age_range}"
                        if row_data.get(col_key, "") != category:
                            is_category_row = False
                            break
                    if not is_category_row:
                        break

                if is_category_row:
                    continue

            variable_name = row_data['Variable']
            ws.cell(row=row_idx, column=1, value=variable_name).border = thin_border

            # Data for each MetS status and age group
            col_idx = 2
            for mets_status in ["Total", "MetS", "Non-MetS"]:
                for age_range in ["under40", "40-59", "60up"]:
                    col_key = f"{mets_status}_{age_range}"
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_key, ""))
                    cell.border = thin_border
                    col_idx += 1

            # P-values for variable
            if variable_name in p_values_df.index:
                for i, status in enumerate(['Total', 'MetS', 'NonMetS']):
                    p_value = p_values_df.loc[variable_name].get(status, np.nan)
                    p_value_str = read_p(p_value)
                    cell = ws.cell(row=row_idx, column=11 + i, value=p_value_str)
                    cell.border = thin_border

            row_idx += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 38
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws.column_dimensions[col_letter].width = 15

    # Rearrange columns
    move_column(ws, 11, 5)
    move_column(ws, 12, 9)

    # Merge cells
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=9)
    ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=13)

def main(excel_filename, sheet_name='Table S2. Baseline Characteristics'):
    print("=" * 80)
    print("Demographic Table Generation - 12 Groups Analysis")
    print("=" * 80)

    # Load and preprocess data
    print("\n[1/4] Loading and preprocessing data...")
    data_path = '../db/total_only_org.csv'
    df = load_and_prepare_data(data_path)
    all_df, men_df, women_df = preprocess_data(df)

    categorical_cols = list(set(df.columns[1:]) - set(CONTINUOUS_COLS)) + METS_COLS + ['MetS']

    print(f"    Total subjects: {len(all_df)}")
    print(f"    Men: {len(men_df)}")
    print(f"    Women: {len(women_df)}")

    # Generate summary statistics for 12 groups
    print("\n[2/4] Generating summary statistics for 12 groups...")
    results_by_category = {}
    results_by_category["12_Groups"] = generate_summary_12_groups(all_df, categorical_cols)
    print("    Summary statistics generation complete!")

    # Calculate p-values for 12 groups
    print("\n[3/4] Calculating p-values for 12 groups...")
    df_clean = all_df.replace('Missing value', np.nan)

    twelve_groups_p_values_dict = {}
    for col in list(df_clean.columns[1:]):
        p_value = calculate_12groups_p_value(df_clean, col, categorical_cols)
        twelve_groups_p_values_dict[col] = p_value

    twelve_groups_p_values = pd.DataFrame({
        'Variable': list(twelve_groups_p_values_dict.keys()),
        'p_value': list(twelve_groups_p_values_dict.values())
    })

    significant_count = sum(twelve_groups_p_values['p_value'] < 0.05)
    print(f"    Variables tested: {len(twelve_groups_p_values)}")
    print(f"    Significant (p<0.05): {significant_count}")

    # Generate Excel file
    print("\n[4/4] Generating Excel file...")

    # Check if file exists
    if os.path.exists(excel_filename):
        print(f"    Loading existing file: {excel_filename}")
        wb = load_workbook(excel_filename)
    else:
        print(f"    Creating new file: {excel_filename}")
        wb = Workbook()
        # Remove default sheet if it exists
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    create_12groups_sheet(wb, results_by_category, twelve_groups_p_values, df_clean, sheet_name)

    wb.save(excel_filename)

    print(f"    Sheet '{sheet_name}' created successfully!")

    print(f"\n{'=' * 80}")
    print(f"Excel file saved successfully: {excel_filename}")
    print(f"Sheets created: {wb.sheetnames}")
    print(f"{'=' * 80}\n")

if __name__ == "__main__":
    # Set paths
    BASE_DIR = Path('../')
    TABLES_DIR = BASE_DIR / 'result' / 'tables'
    TABLES_DIR.mkdir(parents=True, exist_ok=True)

    excel_filename = TABLES_DIR / 'Tables.xlsx'
    sheet_name = 'Table S2. Baseline Characteristics'

    main(str(excel_filename), sheet_name)
